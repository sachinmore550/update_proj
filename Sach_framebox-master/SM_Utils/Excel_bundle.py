from openpyxl import load_workbook


class Excel(object):

    def get_test_data_as_dictionary(self, testdata_filename, sheet_name, testcase_name):

        wb = load_workbook('D://Sachin_Sandbox//Sach_framebox-master//Sach_framebox-master//SM_Testdata//'+testdata_filename)
        ws = wb[sheet_name]
        columns = []
        values = []
        # get headers
        for i in range (1,40):
            if ws.cell(1, i).value:
                columns.append(ws.cell(1, i).value)

        # get row number
        for i in range(1,40):
            if ws.cell(i,2).value == testcase_name:
                x = i
                break

        for i in range (1,len(columns)+1):
            values.append(ws.cell(x, i).value)

        test_data_dict = {key: value for (key, value) in zip(columns,values)}
        return test_data_dict

